import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


def salvar_relatorio_excel(
    df,
    resumo,
    palavras_chave,
    faixa_preco,
    score_info,
    resumo_executivo,
    preco_info,
    resumo_conc,
    insights,
    nome_arquivo="data/relatorio_marketplace.xlsx"
):
    df_resumo = pd.DataFrame([resumo])

    df_palavras = pd.DataFrame(
        palavras_chave,
        columns=["palavra", "frequencia"]
    )

    df_faixa = pd.DataFrame(
        list(faixa_preco.items()),
        columns=["faixa", "quantidade"]
    )

    df_score = pd.DataFrame([score_info])
    df_preco = pd.DataFrame([preco_info])

    df_resumo_executivo = pd.DataFrame(
        [{"resumo_executivo": resumo_executivo}]
    )

    df_concorrentes = pd.DataFrame(
        list(resumo_conc.items()),
        columns=["tipo_concorrente", "quantidade"]
    )

    df_insights = pd.DataFrame(
        [{"insight": insight} for insight in insights]
    )

    with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Produtos", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_palavras.to_excel(writer, sheet_name="Palavras-chave", index=False)
        df_faixa.to_excel(writer, sheet_name="Faixa de Preço", index=False)
        df_score.to_excel(writer, sheet_name="Score", index=False)
        df_preco.to_excel(writer, sheet_name="Preço Sugerido", index=False)
        df_resumo_executivo.to_excel(writer, sheet_name="Resumo Executivo", index=False)
        df_concorrentes.to_excel(writer, sheet_name="Concorrentes", index=False)
        df_insights.to_excel(writer, sheet_name="Insights", index=False)

        cor_titulo = "1F4E78"
        cor_card_azul = "D9EAF7"
        cor_card_verde = "DFF2BF"
        cor_card_amarelo = "FFF2CC"
        cor_card_vermelho = "F4CCCC"
        cor_card_laranja = "FCE5CD"

        for sheet in writer.book.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet.column_dimensions["A"].width = 28
            sheet.column_dimensions["B"].width = 22
            sheet.column_dimensions["C"].width = 22
            sheet.column_dimensions["D"].width = 22

            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=cor_titulo)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_resumo = writer.book["Resumo"]
        ws_resumo.column_dimensions["A"].width = 28
        ws_resumo.column_dimensions["B"].width = 28

        ws_resumo["A1"].font = Font(bold=True, color="FFFFFF")
        ws_resumo["A1"].fill = PatternFill("solid", fgColor=cor_titulo)
        ws_resumo["B1"].font = Font(bold=True, color="FFFFFF")
        ws_resumo["B1"].fill = PatternFill("solid", fgColor=cor_titulo)

        resumo_items = list(resumo.items())
        ws_resumo.delete_rows(2, ws_resumo.max_row)

        for i, (chave, valor) in enumerate(resumo_items, start=2):
            ws_resumo[f"A{i}"] = chave
            ws_resumo[f"B{i}"] = valor

            ws_resumo[f"A{i}"].fill = PatternFill("solid", fgColor=cor_card_azul)
            ws_resumo[f"B{i}"].fill = PatternFill("solid", fgColor=cor_card_verde)

            ws_resumo[f"A{i}"].font = Font(bold=True)
            ws_resumo[f"B{i}"].font = Font(bold=True)

            ws_resumo[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
            ws_resumo[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")

        ws_score = writer.book["Score"]
        ws_score.column_dimensions["A"].width = 28
        ws_score.column_dimensions["B"].width = 28

        nivel = score_info["nivel_oportunidade"].lower()
        cor_score = cor_card_vermelho

        if "alta" in nivel:
            cor_score = cor_card_verde
        elif "média" in nivel or "media" in nivel:
            cor_score = cor_card_amarelo

        score_items = list(score_info.items())
        ws_score.delete_rows(2, ws_score.max_row)

        for i, (chave, valor) in enumerate(score_items, start=2):
            ws_score[f"A{i}"] = chave
            ws_score[f"B{i}"] = valor

            ws_score[f"A{i}"].fill = PatternFill("solid", fgColor=cor_card_azul)
            ws_score[f"B{i}"].fill = PatternFill("solid", fgColor=cor_score)

            ws_score[f"A{i}"].font = Font(bold=True)
            ws_score[f"B{i}"].font = Font(bold=True)

            ws_score[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
            ws_score[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")

        ws_score["D1"] = "categoria"
        ws_score["E1"] = "valor"

        ws_score["D1"].font = Font(bold=True, color="FFFFFF")
        ws_score["D1"].fill = PatternFill("solid", fgColor=cor_titulo)
        ws_score["E1"].font = Font(bold=True, color="FFFFFF")
        ws_score["E1"].fill = PatternFill("solid", fgColor=cor_titulo)

        score_valor = score_info["score"]
        restante = max(10 - score_valor, 0)

        ws_score["D2"] = "Score"
        ws_score["E2"] = score_valor
        ws_score["D3"] = "Restante"
        ws_score["E3"] = restante

        ws_preco = writer.book["Preço Sugerido"]
        ws_preco.column_dimensions["A"].width = 28
        ws_preco.column_dimensions["B"].width = 28

        preco_items = list(preco_info.items())
        ws_preco.delete_rows(2, ws_preco.max_row)

        for i, (chave, valor) in enumerate(preco_items, start=2):
            ws_preco[f"A{i}"] = chave
            ws_preco[f"B{i}"] = valor

            ws_preco[f"A{i}"].fill = PatternFill("solid", fgColor=cor_card_laranja)
            ws_preco[f"B{i}"].fill = PatternFill("solid", fgColor=cor_card_azul)

            ws_preco[f"A{i}"].font = Font(bold=True)
            ws_preco[f"B{i}"].font = Font(bold=True)

            ws_preco[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
            ws_preco[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet_resumo = writer.book["Resumo Executivo"]
        worksheet_resumo.column_dimensions["A"].width = 120
        worksheet_resumo["A1"].font = Font(bold=True, color="FFFFFF")
        worksheet_resumo["A1"].fill = PatternFill("solid", fgColor=cor_titulo)

        cell = worksheet_resumo["A2"]
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        worksheet_insights = writer.book["Insights"]
        worksheet_insights.column_dimensions["A"].width = 120

        for row in range(2, len(df_insights) + 2):
            worksheet_insights[f"A{row}"].alignment = Alignment(wrap_text=True)
            worksheet_insights[f"A{row}"].fill = PatternFill("solid", fgColor=cor_card_amarelo)

        ws_faixa = writer.book["Faixa de Preço"]

        grafico_faixa = BarChart()
        grafico_faixa.title = "Distribuição de Preços"
        grafico_faixa.style = 10
        grafico_faixa.y_axis.title = "Quantidade"
        grafico_faixa.x_axis.title = "Faixa"
        grafico_faixa.height = 7
        grafico_faixa.width = 12

        dados_faixa = Reference(
            ws_faixa,
            min_col=2,
            max_col=2,
            min_row=1,
            max_row=len(df_faixa) + 1
        )

        categorias_faixa = Reference(
            ws_faixa,
            min_col=1,
            min_row=2,
            max_row=len(df_faixa) + 1
        )

        grafico_faixa.add_data(dados_faixa, titles_from_data=True)
        grafico_faixa.set_categories(categorias_faixa)

        grafico_faixa.dataLabels = DataLabelList()
        grafico_faixa.dataLabels.showVal = True

        ws_faixa.add_chart(grafico_faixa, "E2")

        grafico_score = DoughnutChart()
        grafico_score.title = "Score de Oportunidade"
        grafico_score.style = 10
        grafico_score.holeSize = 60
        grafico_score.height = 7
        grafico_score.width = 10

        dados_score = Reference(
            ws_score,
            min_col=5,
            max_col=5,
            min_row=1,
            max_row=3
        )

        categorias_score = Reference(
            ws_score,
            min_col=4,
            min_row=2,
            max_row=3
        )

        grafico_score.add_data(dados_score, titles_from_data=True)
        grafico_score.set_categories(categorias_score)

        ws_score.add_chart(grafico_score, "G2")

        ws_palavras = writer.book["Palavras-chave"]

        grafico_palavras = BarChart()
        grafico_palavras.title = "Palavras-Chave Mais Frequentes"
        grafico_palavras.style = 11
        grafico_palavras.y_axis.title = "Frequência"
        grafico_palavras.x_axis.title = "Palavra"
        grafico_palavras.height = 8
        grafico_palavras.width = 14

        dados_palavras = Reference(
            ws_palavras,
            min_col=2,
            max_col=2,
            min_row=1,
            max_row=len(df_palavras) + 1
        )

        categorias_palavras = Reference(
            ws_palavras,
            min_col=1,
            min_row=2,
            max_row=len(df_palavras) + 1
        )

        grafico_palavras.add_data(dados_palavras, titles_from_data=True)
        grafico_palavras.set_categories(categorias_palavras)

        grafico_palavras.dataLabels = DataLabelList()
        grafico_palavras.dataLabels.showVal = True

        ws_palavras.add_chart(grafico_palavras, "D2")

        ws_concorrentes = writer.book["Concorrentes"]

        grafico_concorrentes = PieChart()
        grafico_concorrentes.title = "Tipos de Concorrentes"
        grafico_concorrentes.style = 10
        grafico_concorrentes.height = 8
        grafico_concorrentes.width = 10

        dados_concorrentes = Reference(
            ws_concorrentes,
            min_col=2,
            max_col=2,
            min_row=1,
            max_row=len(df_concorrentes) + 1
        )

        categorias_concorrentes = Reference(
            ws_concorrentes,
            min_col=1,
            min_row=2,
            max_row=len(df_concorrentes) + 1
        )

        grafico_concorrentes.add_data(dados_concorrentes, titles_from_data=True)
        grafico_concorrentes.set_categories(categorias_concorrentes)

        ws_concorrentes.add_chart(grafico_concorrentes, "D2")